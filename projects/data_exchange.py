from __future__ import annotations

import csv
import io
import json
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from projects.project_manager import append_project_history
from projects.repository import safe_project_id
from projects.well_cards import safe_well_id

PROJECT_DATA_EXCHANGE_FILE_NAME = "data_exchange.json"
DATA_EXCHANGE_FORMATS = {"csv", "xlsx", "json", "geojson", "dlis", "lis", "project_zip"}
DATA_EXCHANGE_DIRECTIONS = {"import", "export"}
DATA_EXCHANGE_STATUSES = {"planned", "ready", "done", "failed", "skipped"}
TABULAR_FORMATS = {"csv", "xlsx", "json"}


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _project_dir(root: Path | str, project_id: str) -> Path:
    return Path(root) / safe_project_id(project_id)


def _exchange_path(root: Path | str, project_id: str) -> Path:
    return _project_dir(root, project_id) / PROJECT_DATA_EXCHANGE_FILE_NAME


def _json_read(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError, TypeError):
        return default


def _json_write(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _payload(root: Path | str, project_id: str) -> dict[str, Any]:
    payload = _json_read(_exchange_path(root, project_id), {"records": [], "profiles": []})
    if not isinstance(payload, dict):
        payload = {"records": [], "profiles": []}
    payload.setdefault("records", [])
    payload.setdefault("profiles", [])
    return payload


def _clean_text(value: Any, field_label: str, *, max_length: int = 240, required: bool = False) -> str:
    text = "" if value is None else str(value).strip()
    if required and not text:
        raise ValueError(f"{field_label}: значение обязательно.")
    if len(text) > max_length:
        raise ValueError(f"{field_label}: максимум {max_length} символов.")
    return text


def _clean_format(value: Any) -> str:
    fmt = _clean_text(value, "Формат", max_length=40, required=True).lower().lstrip(".")
    if fmt not in DATA_EXCHANGE_FORMATS:
        raise ValueError(f"Формат должен быть одним из: {', '.join(sorted(DATA_EXCHANGE_FORMATS))}.")
    return fmt


def _clean_direction(value: Any) -> str:
    direction = _clean_text(value, "Направление", max_length=20).lower() or "import"
    if direction not in DATA_EXCHANGE_DIRECTIONS:
        raise ValueError(f"Направление должно быть одним из: {', '.join(sorted(DATA_EXCHANGE_DIRECTIONS))}.")
    return direction


def _clean_status(value: Any) -> str:
    status = _clean_text(value, "Статус", max_length=20).lower() or "planned"
    if status not in DATA_EXCHANGE_STATUSES:
        raise ValueError(f"Статус должен быть одним из: {', '.join(sorted(DATA_EXCHANGE_STATUSES))}.")
    return status


def _safe_record_id(value: Any, *, default: str = "exchange") -> str:
    text = _clean_text(value, "ID", max_length=160) or default
    normalized = "".join(ch if ch.isalnum() or ch in "_-" else "-" for ch in text).strip("-_").lower()
    return safe_well_id(normalized or default)


@dataclass(frozen=True)
class DataExchangeRecord:
    """Single import/export operation registered in the project exchange log."""

    id: str
    name: str
    direction: str
    format: str
    source_path: str = ""
    target_path: str = ""
    status: str = "planned"
    rows: int = 0
    columns: int = 0
    warnings: tuple[str, ...] = ()
    errors: tuple[str, ...] = ()
    metadata: Mapping[str, Any] = field(default_factory=dict)
    created_at: str = ""
    updated_at: str = ""


@dataclass(frozen=True)
class ExchangeValidationIssue:
    severity: str
    code: str
    message: str
    field: str = ""
    recommendation: str = ""


@dataclass(frozen=True)
class DataExchangeProfile:
    id: str
    name: str
    format: str
    direction: str = "import"
    delimiter: str = ","
    encoding: str = "utf-8"
    column_mapping: Mapping[str, str] = field(default_factory=dict)
    required_columns: tuple[str, ...] = ()
    metadata: Mapping[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class DataExchangeSummary:
    records: int
    imports: int
    exports: int
    failed: int
    formats: tuple[str, ...]
    rows: int


def normalize_exchange_record(raw: DataExchangeRecord | Mapping[str, Any]) -> DataExchangeRecord:
    if isinstance(raw, DataExchangeRecord):
        row = raw
    elif isinstance(raw, Mapping):
        now = _utc_now()
        row = DataExchangeRecord(
            id=_safe_record_id(raw.get("id") or raw.get("name") or raw.get("source_path") or raw.get("target_path")),
            name=_clean_text(raw.get("name") or raw.get("source_path") or raw.get("target_path"), "Название", required=True),
            direction=_clean_direction(raw.get("direction", "import")),
            format=_clean_format(raw.get("format") or Path(str(raw.get("source_path") or raw.get("target_path") or "")).suffix),
            source_path=_clean_text(raw.get("source_path"), "Источник", max_length=500),
            target_path=_clean_text(raw.get("target_path"), "Назначение", max_length=500),
            status=_clean_status(raw.get("status", "planned")),
            rows=max(0, int(raw.get("rows") or 0)),
            columns=max(0, int(raw.get("columns") or 0)),
            warnings=tuple(str(item) for item in raw.get("warnings", ()) if str(item).strip()),
            errors=tuple(str(item) for item in raw.get("errors", ()) if str(item).strip()),
            metadata=raw.get("metadata", {}) if isinstance(raw.get("metadata", {}), Mapping) else {},
            created_at=_clean_text(raw.get("created_at") or now, "Дата создания", max_length=80),
            updated_at=_clean_text(raw.get("updated_at") or now, "Дата обновления", max_length=80),
        )
    else:
        raise TypeError("Exchange record должен быть DataExchangeRecord или mapping.")
    return DataExchangeRecord(
        id=_safe_record_id(row.id),
        name=_clean_text(row.name, "Название", required=True),
        direction=_clean_direction(row.direction),
        format=_clean_format(row.format),
        source_path=_clean_text(row.source_path, "Источник", max_length=500),
        target_path=_clean_text(row.target_path, "Назначение", max_length=500),
        status=_clean_status(row.status),
        rows=max(0, int(row.rows or 0)),
        columns=max(0, int(row.columns or 0)),
        warnings=tuple(str(item).strip() for item in row.warnings if str(item).strip()),
        errors=tuple(str(item).strip() for item in row.errors if str(item).strip()),
        metadata=dict(row.metadata),
        created_at=row.created_at or _utc_now(),
        updated_at=row.updated_at or _utc_now(),
    )


def _record_to_dict(record: DataExchangeRecord) -> dict[str, Any]:
    return {
        "id": record.id,
        "name": record.name,
        "direction": record.direction,
        "format": record.format,
        "source_path": record.source_path,
        "target_path": record.target_path,
        "status": record.status,
        "rows": record.rows,
        "columns": record.columns,
        "warnings": list(record.warnings),
        "errors": list(record.errors),
        "metadata": dict(record.metadata),
        "created_at": record.created_at,
        "updated_at": record.updated_at,
    }


def list_data_exchange_records(root: Path | str, project_id: str, *, direction: str = "", format: str = "") -> tuple[DataExchangeRecord, ...]:
    payload = _payload(root, project_id)
    records = []
    for raw in payload.get("records", []):
        if not isinstance(raw, Mapping):
            continue
        try:
            records.append(normalize_exchange_record(raw))
        except (TypeError, ValueError):
            continue
    if direction:
        clean_direction = _clean_direction(direction)
        records = [record for record in records if record.direction == clean_direction]
    if format:
        clean_format = _clean_format(format)
        records = [record for record in records if record.format == clean_format]
    return tuple(sorted(records, key=lambda row: (row.updated_at, row.id), reverse=True))


def save_data_exchange_record(root: Path | str, project_id: str, record: DataExchangeRecord | Mapping[str, Any]) -> DataExchangeRecord:
    normalized = normalize_exchange_record(record)
    payload = _payload(root, project_id)
    kept = [raw for raw in payload.get("records", []) if isinstance(raw, Mapping) and str(raw.get("id")) != normalized.id]
    payload["records"] = [_record_to_dict(normalized), *kept]
    _json_write(_exchange_path(root, project_id), {"version": 1, **payload})
    append_project_history(root, project_id, "data-exchange", f"Saved {normalized.direction} {normalized.format}: {normalized.name}", object_type="data_exchange", object_id=normalized.id)
    return normalized


def delete_data_exchange_record(root: Path | str, project_id: str, record_id: str) -> bool:
    clean_id = _safe_record_id(record_id)
    payload = _payload(root, project_id)
    kept = [raw for raw in payload.get("records", []) if isinstance(raw, Mapping) and str(raw.get("id")) != clean_id]
    if len(kept) == len(payload.get("records", [])):
        return False
    payload["records"] = kept
    _json_write(_exchange_path(root, project_id), {"version": 1, **payload})
    append_project_history(root, project_id, "data-exchange", f"Deleted exchange record {clean_id}", object_type="data_exchange", object_id=clean_id)
    return True


def normalize_exchange_profile(raw: DataExchangeProfile | Mapping[str, Any]) -> DataExchangeProfile:
    if isinstance(raw, DataExchangeProfile):
        profile = raw
    elif isinstance(raw, Mapping):
        profile = DataExchangeProfile(
            id=_safe_record_id(raw.get("id") or raw.get("name"), default="profile"),
            name=_clean_text(raw.get("name"), "Название профиля", required=True),
            format=_clean_format(raw.get("format")),
            direction=_clean_direction(raw.get("direction", "import")),
            delimiter=_clean_text(raw.get("delimiter", ","), "Разделитель", max_length=8) or ",",
            encoding=_clean_text(raw.get("encoding", "utf-8"), "Кодировка", max_length=40) or "utf-8",
            column_mapping=raw.get("column_mapping", {}) if isinstance(raw.get("column_mapping", {}), Mapping) else {},
            required_columns=tuple(str(item).strip() for item in raw.get("required_columns", ()) if str(item).strip()),
            metadata=raw.get("metadata", {}) if isinstance(raw.get("metadata", {}), Mapping) else {},
        )
    else:
        raise TypeError("Exchange profile должен быть DataExchangeProfile или mapping.")
    return DataExchangeProfile(
        id=_safe_record_id(profile.id, default="profile"),
        name=_clean_text(profile.name, "Название профиля", required=True),
        format=_clean_format(profile.format),
        direction=_clean_direction(profile.direction),
        delimiter=_clean_text(profile.delimiter, "Разделитель", max_length=8) or ",",
        encoding=_clean_text(profile.encoding, "Кодировка", max_length=40) or "utf-8",
        column_mapping={str(k): str(v) for k, v in dict(profile.column_mapping).items()},
        required_columns=tuple(str(item).strip() for item in profile.required_columns if str(item).strip()),
        metadata=dict(profile.metadata),
    )


def validate_exchange_table(rows: Sequence[Mapping[str, Any]], *, required_columns: Iterable[str] = ()) -> tuple[ExchangeValidationIssue, ...]:
    issues: list[ExchangeValidationIssue] = []
    if not rows:
        issues.append(ExchangeValidationIssue("warning", "empty-table", "Таблица не содержит строк.", recommendation="Проверьте исходный файл или диапазон импорта."))
        return tuple(issues)
    columns = set(rows[0].keys())
    for column in required_columns:
        if column not in columns:
            issues.append(ExchangeValidationIssue("error", "missing-column", f"Отсутствует обязательная колонка: {column}.", field=column, recommendation="Добавьте колонку или настройте mapping профиля."))
    inconsistent = [index + 1 for index, row in enumerate(rows) if set(row.keys()) != columns]
    if inconsistent:
        issues.append(ExchangeValidationIssue("warning", "inconsistent-columns", f"Строки с отличающимся набором колонок: {', '.join(map(str, inconsistent[:8]))}.", recommendation="Нормализуйте структуру таблицы перед импортом."))
    return tuple(issues)


def import_csv_text(text: str, *, delimiter: str = ",", required_columns: Iterable[str] = ()) -> tuple[dict[str, str], ...]:
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = tuple(dict(row) for row in reader)
    errors = [issue.message for issue in validate_exchange_table(rows, required_columns=required_columns) if issue.severity == "error"]
    if errors:
        raise ValueError("; ".join(errors))
    return rows


def export_rows_csv(rows: Sequence[Mapping[str, Any]], *, delimiter: str = ",") -> str:
    output = io.StringIO()
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(str(key))
    writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=delimiter, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in fieldnames})
    return output.getvalue()


def import_json_text(text: str, *, required_columns: Iterable[str] = ()) -> tuple[dict[str, Any], ...]:
    data = json.loads(text)
    if isinstance(data, Mapping):
        data = data.get("rows", data.get("features", []))
    rows: list[dict[str, Any]] = []
    for item in data if isinstance(data, list) else []:
        if isinstance(item, Mapping) and item.get("type") == "Feature":
            props = item.get("properties", {}) if isinstance(item.get("properties", {}), Mapping) else {}
            rows.append({**props, "geometry": item.get("geometry")})
        elif isinstance(item, Mapping):
            rows.append(dict(item))
    errors = [issue.message for issue in validate_exchange_table(rows, required_columns=required_columns) if issue.severity == "error"]
    if errors:
        raise ValueError("; ".join(errors))
    return tuple(rows)


def export_rows_json(rows: Sequence[Mapping[str, Any]]) -> str:
    return json.dumps({"schema": "gas-ratio-pro.rows.v1", "rows": [dict(row) for row in rows]}, ensure_ascii=False, indent=2)


def export_rows_geojson(rows: Sequence[Mapping[str, Any]], *, lon_column: str = "lon", lat_column: str = "lat") -> str:
    features = []
    for row in rows:
        lon = row.get(lon_column)
        lat = row.get(lat_column)
        try:
            point = [float(lon), float(lat)]
        except (TypeError, ValueError):
            continue
        properties = {str(k): v for k, v in row.items() if k not in {lon_column, lat_column}}
        features.append({"type": "Feature", "geometry": {"type": "Point", "coordinates": point}, "properties": properties})
    return json.dumps({"type": "FeatureCollection", "features": features}, ensure_ascii=False, indent=2)


def build_xlsx_bytes(rows: Sequence[Mapping[str, Any]], *, sheet_name: str = "Data") -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - depends on optional dependency
        raise RuntimeError("Для XLSX-экспорта требуется пакет openpyxl.") from exc
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(str(key))
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Data"
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(key, "") for key in fieldnames])
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def read_xlsx_bytes(data: bytes, *, sheet_name: str | None = None) -> tuple[dict[str, Any], ...]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Для XLSX-импорта требуется пакет openpyxl.") from exc
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter, ())]
    result = []
    for values in rows_iter:
        result.append({headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]})
    return tuple(result)


def build_project_exchange_manifest(root: Path | str, project_id: str) -> dict[str, Any]:
    project_path = _project_dir(root, project_id)
    files = []
    for path in sorted(project_path.rglob("*")):
        if path.is_file():
            files.append({"path": path.relative_to(project_path).as_posix(), "size_bytes": path.stat().st_size})
    return {"schema": "gas-ratio-pro.project-exchange.v1", "project_id": safe_project_id(project_id), "created_at": _utc_now(), "files": files}


def build_project_exchange_zip(root: Path | str, project_id: str, *, include_patterns: Iterable[str] = ()) -> bytes:
    project_path = _project_dir(root, project_id)
    patterns = tuple(include_patterns)
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        manifest = build_project_exchange_manifest(root, project_id)
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        for path in sorted(project_path.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(project_path).as_posix()
            if patterns and not any(Path(rel).match(pattern) for pattern in patterns):
                continue
            archive.write(path, rel)
    return stream.getvalue()


def summarize_data_exchange(records: Iterable[DataExchangeRecord]) -> DataExchangeSummary:
    items = tuple(records)
    return DataExchangeSummary(
        records=len(items),
        imports=sum(1 for item in items if item.direction == "import"),
        exports=sum(1 for item in items if item.direction == "export"),
        failed=sum(1 for item in items if item.status == "failed"),
        formats=tuple(sorted({item.format for item in items})),
        rows=sum(item.rows for item in items),
    )


def build_data_exchange_record_table(records: Iterable[DataExchangeRecord]) -> tuple[dict[str, Any], ...]:
    return tuple(
        {
            "ID": record.id,
            "Название": record.name,
            "Направление": record.direction,
            "Формат": record.format.upper(),
            "Статус": record.status,
            "Строк": record.rows,
            "Колонок": record.columns,
            "Источник": record.source_path or "—",
            "Назначение": record.target_path or "—",
            "Предупреждения": len(record.warnings),
            "Ошибки": len(record.errors),
            "Обновлено": record.updated_at,
        }
        for record in records
    )


def build_exchange_issue_table(issues: Iterable[ExchangeValidationIssue]) -> tuple[dict[str, str], ...]:
    return tuple(
        {
            "Уровень": issue.severity,
            "Код": issue.code,
            "Поле": issue.field or "—",
            "Сообщение": issue.message,
            "Рекомендация": issue.recommendation or "—",
        }
        for issue in issues
    )
