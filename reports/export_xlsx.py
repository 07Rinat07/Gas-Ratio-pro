from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_sheet_name(value: str, fallback: str) -> str:
    cleaned = "".join("_" if char in r'[]:*?/\\' else char for char in str(value)).strip()
    return (cleaned or fallback)[:31]


def _fit_columns(worksheet, *, max_width: int = 48) -> None:
    for column_cells in worksheet.columns:
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, max((len(line) for line in value.splitlines()), default=0))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), max_width)


def export_xlsx_bytes(
    df: pd.DataFrame,
    sheet_name: str = "calculations",
    *,
    metadata: Mapping[str, Any] | None = None,
) -> bytes:
    """Create a readable XLSX export with optional report metadata.

    The metadata sheet makes the selected profile and depth range explicit and
    prevents an XLSX file from being detached from the settings used to create
    it.  The data sheet still contains the complete filtered engineering frame.
    """
    if df is None or df.empty:
        return b""

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        data_sheet = _safe_sheet_name(sheet_name, "calculations")
        df.to_excel(writer, index=False, sheet_name=data_sheet)
        workbook = writer.book
        data_ws = writer.sheets[data_sheet]
        data_ws.freeze_panes = "A2"
        data_ws.auto_filter.ref = data_ws.dimensions
        for cell in data_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _fit_columns(data_ws)

        if metadata:
            meta_sheet = _safe_sheet_name("Параметры отчёта", "report_metadata")
            meta_ws = workbook.create_sheet(meta_sheet, 0)
            meta_ws.append(["Параметр", "Значение"])
            for key, value in metadata.items():
                meta_ws.append([str(key), "" if value is None else str(value)])
            for cell in meta_ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for row in meta_ws.iter_rows(min_row=2, min_col=1, max_col=2):
                row[0].font = Font(bold=True)
                row[1].alignment = Alignment(wrap_text=True, vertical="top")
            meta_ws.freeze_panes = "A2"
            _fit_columns(meta_ws, max_width=64)

    return buffer.getvalue()
