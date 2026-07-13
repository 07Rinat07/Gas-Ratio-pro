from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from reports.export_controller import (
    ExportArtifact,
    ExportController,
    ExportControllerError,
    ExportRequest,
)
from reports.export_xlsx import export_xlsx_bytes


def _request(**overrides) -> ExportRequest:
    values = dict(
        project_id="default",
        project_name="Основной проект",
        source_label="LAS",
        profile_id="client",
        format_id="pdf",
        format_label="PDF",
        extension="pdf",
        mime_type="application/pdf",
        depth_top=1000.0,
        depth_bottom=1010.0,
        source_signature="source-1",
        calculation_revision=2,
        presentation_revision=3,
        figure_height=1200,
        context_signature="context-1",
    )
    values.update(overrides)
    return ExportRequest(**values)


def test_selection_signature_changes_for_user_visible_choices() -> None:
    base = _request()
    assert base.selection_signature != _request(profile_id="engineering").selection_signature
    assert base.selection_signature != _request(format_id="docx", extension="docx", mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document").selection_signature
    assert base.selection_signature != _request(depth_bottom=1011.0).selection_signature


def test_controller_rejects_renderer_profile_mismatch() -> None:
    controller = ExportController({})

    with pytest.raises(ExportControllerError) as error:
        controller.prepare(
            _request(),
            frame=pd.DataFrame({"depth": [1000.0]}),
            build_model=lambda frame, request: object(),
            render_artifact=lambda model, frame, request: ExportArtifact(
                content=b"%PDF-1.4",
                file_name="report.pdf",
                mime_type="application/pdf",
                format_id="pdf",
                format_label="PDF",
                profile_id="engineering",
            ),
        )

    assert error.value.failure.stage == "render_pdf"
    assert "profile_id" in error.value.failure.message


def test_controller_stores_request_signature_on_artifact() -> None:
    request = _request()
    artifact, _ = ExportController({}).prepare(
        request,
        frame=pd.DataFrame({"depth": [1000.0]}),
        build_model=lambda frame, export_request: object(),
        render_artifact=lambda model, frame, export_request: ExportArtifact(
            content=b"%PDF-1.4",
            file_name="report.pdf",
            mime_type="application/pdf",
            format_id="pdf",
            format_label="PDF",
            profile_id="client",
        ),
    )
    assert artifact.request_signature == request.selection_signature


def test_xlsx_contains_metadata_and_engineering_data_sheets() -> None:
    content = export_xlsx_bytes(
        pd.DataFrame({"depth": [1000.0, 1000.2], "c1": [1.0, 2.0]}),
        sheet_name="Инженерные данные",
        metadata={
            "Профиль отчёта": "Для заказчика",
            "Глубина от, м": 1000,
            "Глубина до, м": 1000.2,
        },
    )
    workbook = load_workbook(BytesIO(content), read_only=False)
    assert workbook.sheetnames[0] == "Параметры отчёта"
    assert "Инженерные данные" in workbook.sheetnames
    metadata_values = {
        workbook["Параметры отчёта"].cell(row=row, column=1).value:
        workbook["Параметры отчёта"].cell(row=row, column=2).value
        for row in range(2, workbook["Параметры отчёта"].max_row + 1)
    }
    assert metadata_values["Профиль отчёта"] == "Для заказчика"
    assert workbook["Инженерные данные"].freeze_panes == "A2"
